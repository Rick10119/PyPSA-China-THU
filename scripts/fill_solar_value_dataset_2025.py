#!/usr/bin/env python3
"""
Fill solar value dataset rows for planning years 2025-2060.

Inputs per year:
- dispatch segmented network (.nc): solar generation/demand/capacity side metrics
- provincial LMP time series for value-factor metrics (default: planning postnetwork
  `buses_t.marginal_price`, EUR/MWh converted with FX to match legacy CSV convention)
- optional: mapped dispatch prices CSV (pass --mapped-csv)

Capacity adjustment rule:
- 2025: apply real-capacity correction from `solar capacity.csv`
- 2030 and later: no capacity correction (use model capacity directly)
"""

from __future__ import annotations

import argparse
from pathlib import Path
from shutil import copy2

import sys

import pandas as pd
import pypsa
from openpyxl import load_workbook

_SCRIPTS = Path(__file__).resolve().parent
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))

from reconstruct_market_prices import ReconstructPriceConfig, marginal_retail_prices  # noqa: E402

ROOT = Path(__file__).resolve().parents[1]
VERSION_DIR = ROOT / "results" / "version-0509.1H.1"
HEATING_DEMAND = "positive"
SCENARIO_STEM = "ll-current+FCG-linear2050"
TARGET_YEARS = list(range(2025, 2065, 5))
XLSX_PATH = VERSION_DIR / "solar_value_dataset.xlsx"
BACKUP_PATH = VERSION_DIR / "solar_value_dataset.multi-year-backup.xlsx"
REAL_SOLAR_CAP_PATH = ROOT / "data" / "existing_infrastructure" / "solar capacity.csv"
CAP_COMPARE_PATH = VERSION_DIR / "solar_capacity_compare_by_year.csv"

# Default: LMPs from capacity-planning solve (EUR/MWh in typical PyPSA cost tables).
PLANNING_LMP_FX_CNY_PER_EUR = 7.8

# Workbook province -> source province names
PROVINCE_MAP = {
    "Xizang": "Tibet",
    "WestInnerMongolia": "InnerMongolia",
    "EastInnerMongolia": "InnerMongolia",
}

# Proxy split for Inner Mongolia east/west (applied to all modeled years).
# Source used: 2025 Inner Mongolia power bulletin reports generation shares
# (蒙东 21.93%, 蒙西 78.07%). Public direct PV east/west split was not found.
INNER_MONGOLIA_SPLIT_2025 = {
    "EastInnerMongolia": 0.2193,
    "WestInnerMongolia": 0.7807,
}


def _mapped_name(province: str) -> str:
    return PROVINCE_MAP.get(province, province)


def _group_sum_by_bus(frame: pd.DataFrame, bus_of_component: pd.Series) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame(index=frame.index)
    return frame.T.groupby(bus_of_component).sum().T


def _network_path(year: int) -> Path:
    return (
        VERSION_DIR
        / "dispatch_segmented"
        / HEATING_DEMAND
        / f"postnetwork-dispatch-seg-{SCENARIO_STEM}-{year}.nc"
    )


def _planning_network_path(year: int) -> Path:
    return (
        VERSION_DIR
        / "postnetworks"
        / HEATING_DEMAND
        / f"postnetwork-{SCENARIO_STEM}-{year}.nc"
    )


def _price_column_for_bus(bus: str, price_cols: set[str]) -> str | None:
    """Resolve dispatch bus name to a column in the planning LMP frame (single InnerMongolia etc.)."""
    if bus in price_cols:
        return bus
    alt = PROVINCE_MAP.get(bus)
    if alt and alt in price_cols:
        return alt
    return None


def _price_csv_path(year: int) -> Path:
    return (
        VERSION_DIR
        / "prices"
        / "dispatch_segmented"
        / HEATING_DEMAND
        / f"dispatch_segmented_prices-{SCENARIO_STEM}-{year}_mapped.csv"
    )


def _block_starts(ws) -> list[int]:
    zones = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
    first_zone = zones[0]
    return [i + 2 for i, z in enumerate(zones) if z == first_zone]


def _load_real_solar_capacity_2025() -> pd.Series:
    cap = pd.read_csv(REAL_SOLAR_CAP_PATH)
    required_cols = ["Region", "2010", "2015", "2020", "2025"]
    if any(c not in cap.columns for c in required_cols):
        raise ValueError(
            "Expected columns 'Region', '2010', '2015', '2020', '2025' in solar capacity.csv"
        )
    real_cap = cap.set_index("Region")[["2010", "2015", "2020", "2025"]].astype(float).sum(axis=1)
    return real_cap


def _compute_metrics_for_year(
    year: int,
    adjust_capacity: bool,
    *,
    price_source: str,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    network_path = _network_path(year)
    if not network_path.exists():
        raise FileNotFoundError(f"Dispatch network not found for {year}: {network_path}")

    n = pypsa.Network(network_path)
    snapshots = pd.DatetimeIndex(n.snapshots)

    if price_source == "planning_marginal":
        planning_path = _planning_network_path(year)
        if not planning_path.exists():
            raise FileNotFoundError(f"Planning network not found for {year}: {planning_path}")
        n_plan = pypsa.Network(planning_path)
        prices = marginal_retail_prices(n_plan, config=ReconstructPriceConfig()).astype(float)
        prices = prices * float(PLANNING_LMP_FX_CNY_PER_EUR)
        prices = prices.reindex(snapshots)
        if prices.isna().any().any():
            raise ValueError(
                f"Planning LMPs for {year} do not cover all dispatch snapshots (check {planning_path})."
            )
    elif price_source == "mapped_csv":
        price_csv_path = _price_csv_path(year)
        if not price_csv_path.exists():
            raise FileNotFoundError(f"Mapped price CSV not found for {year}: {price_csv_path}")
        prices = pd.read_csv(price_csv_path)
        prices = prices.rename(columns={"snapshot": "time"})
        prices["time"] = pd.to_datetime(prices["time"])
        prices = prices.set_index("time").sort_index()
        prices = prices.reindex(snapshots)
        if prices.isna().any().any():
            raise ValueError("Mapped price CSV and network snapshots do not align exactly.")
    else:
        raise ValueError(f"Unknown price_source: {price_source!r}")

    price_col_set = set(prices.columns.astype(str))

    # Select solar generators (include all carriers containing 'solar').
    solar_mask = n.generators.carrier.astype(str).str.contains("solar", case=False, regex=False)
    solar_gens = n.generators.index[solar_mask]

    if len(solar_gens) == 0:
        raise ValueError("No solar generators found in network.")

    solar_dispatch = n.generators_t.p[solar_gens].clip(lower=0.0)
    solar_bus = n.generators.loc[solar_gens, "bus"]
    solar_dispatch_bus = _group_sum_by_bus(solar_dispatch, solar_bus)

    avail = n.generators_t.p_max_pu[solar_gens].multiply(
        n.generators.loc[solar_gens, "p_nom_opt"], axis=1
    )
    solar_available_bus = _group_sum_by_bus(avail, solar_bus)

    solar_capacity_bus = n.generators.loc[solar_gens].groupby("bus")["p_nom_opt"].sum()
    real_solar_capacity_bus = _load_real_solar_capacity_2025() if adjust_capacity else pd.Series(dtype=float)

    ac_buses = n.buses.index[n.buses.carrier.astype(str) == "AC"]

    # System total generation by AC bus/hour:
    # use all generators connected to AC buses, clipped at >=0 to represent generation output.
    all_ac_gen_mask = n.generators.bus.isin(ac_buses)
    all_ac_gens = n.generators.index[all_ac_gen_mask]
    all_gen_dispatch = n.generators_t.p[all_ac_gens].clip(lower=0.0)
    all_gen_bus = n.generators.loc[all_ac_gens, "bus"]
    total_gen_bus = _group_sum_by_bus(all_gen_dispatch, all_gen_bus)

    # Standard value-factor denominator:
    # system weighted-average market value = sum(total_gen * nodal_price) / sum(total_gen)
    system_value_num = 0.0
    system_gen_sum = 0.0
    for b in total_gen_bus.columns:
        pc = _price_column_for_bus(str(b), price_col_set)
        if pc is None:
            continue
        g = total_gen_bus[b]
        p = prices[pc].astype(float)
        system_value_num += float((g * p).sum())
        system_gen_sum += float(g.sum())
    system_avg_market_value = (system_value_num / system_gen_sum) if system_gen_sum > 0 else 0.0

    # Provincial demand on AC buses.
    load_values = n.loads_t.p_set if hasattr(n.loads_t, "p_set") else n.loads_t.p
    load_mask = n.loads.bus.isin(ac_buses)
    load_cols = n.loads.index[load_mask]
    loads = load_values[load_cols].clip(lower=0.0)
    load_bus = n.loads.loc[load_cols, "bus"]
    load_bus_ts = _group_sum_by_bus(loads, load_bus)

    provinces = sorted(set(ac_buses).union(prices.columns))
    weight_sum = float(n.snapshot_weightings.generators.sum())

    cap_compare_rows: list[dict[str, float | str]] = []
    rows: list[dict[str, float | str]] = []
    for province in provinces:
        dispatch_s = solar_dispatch_bus.get(province, pd.Series(0.0, index=snapshots))
        available_s = solar_available_bus.get(province, pd.Series(0.0, index=snapshots))
        total_gen_s = total_gen_bus.get(province, pd.Series(0.0, index=snapshots))
        load_s = load_bus_ts.get(province, pd.Series(0.0, index=snapshots))
        pc = _price_column_for_bus(str(province), price_col_set)
        price_s = prices[pc].astype(float) if pc is not None else pd.Series(0.0, index=snapshots)

        solar_mwh = float(dispatch_s.sum())
        solar_gwh = solar_mwh / 1000.0
        demand_mwh = float(load_s.sum())
        total_gen_mwh = float(total_gen_s.sum())
        available_mwh = float(available_s.sum())
        nc_cap_mw = float(solar_capacity_bus.get(province, 0.0))
        real_cap_mw = float(real_solar_capacity_bus.get(province, nc_cap_mw)) if adjust_capacity else nc_cap_mw
        cap_ratio = (real_cap_mw / nc_cap_mw) if nc_cap_mw > 0 else 1.0
        solar_mwh_for_penetration = solar_mwh * cap_ratio

        pv_value_num = float((dispatch_s * price_s).sum())
        pv_avg_market_value = (pv_value_num / solar_mwh) if solar_mwh > 0 else 0.0
        value_num = pv_avg_market_value
        value_den = system_avg_market_value
        value_factor = (value_num / value_den) if value_den > 0 else 0.0
        # Conservative provincial penetration: take the lower of
        # load-based and generation-based shares.
        penetration_load = (solar_mwh_for_penetration / demand_mwh) if demand_mwh > 0 else 0.0
        penetration_gen = (solar_mwh_for_penetration / total_gen_mwh) if total_gen_mwh > 0 else 0.0
        penetration = min(penetration_load, penetration_gen)
        curtailment = ((available_mwh - solar_mwh) / available_mwh) if available_mwh > 0 else 0.0
        cap_factor = (solar_mwh / (nc_cap_mw * weight_sum)) if nc_cap_mw > 0 and weight_sum > 0 else 0.0

        cap_compare_rows.append(
            {
                "year": year,
                "province": province,
                "nc_solar_capacity_mw": nc_cap_mw,
                "real_solar_capacity_mw": real_cap_mw,
                "real_to_nc_ratio": cap_ratio,
                "capacity_adjusted": adjust_capacity,
            }
        )

        rows.append(
            {
                "province": province,
                "solar_ele_GWh": solar_gwh,
                "value_factor_numerator": value_num,
                "value_factor_denominator": value_den,
                "value_factor": value_factor,
                "solar_penetration": penetration,
                "solar_curtailment_rate": curtailment,
                "solar_capacity_factor": cap_factor,
            }
        )

    cap_compare_df = pd.DataFrame(cap_compare_rows)
    return pd.DataFrame(rows).set_index("province"), cap_compare_df


def _write_workbook(metrics_by_year: dict[int, pd.DataFrame]) -> None:
    copy2(XLSX_PATH, BACKUP_PATH)

    wb = load_workbook(XLSX_PATH)
    ws = wb["Sheet1"]

    starts = _block_starts(ws)
    years_by_block = [2025 + 5 * i for i in range(len(starts))]
    block_size = (starts[1] - starts[0]) if len(starts) > 1 else 32

    for start, year in zip(starts, years_by_block):
        if year not in TARGET_YEARS or year not in metrics_by_year:
            continue
        metrics = metrics_by_year[year]
        for row in range(start, min(start + block_size, ws.max_row + 1)):
            zone = ws.cell(row=row, column=1).value
            if not zone:
                continue
            zone = str(zone)
            source_zone = _mapped_name(zone)
            if source_zone not in metrics.index:
                raise KeyError(f"Province not found in computed metrics: {zone} -> {source_zone}")

            m = metrics.loc[source_zone]
            solar_ele_gwh = float(m["solar_ele_GWh"])
            if source_zone == "InnerMongolia" and zone in INNER_MONGOLIA_SPLIT_2025:
                solar_ele_gwh *= INNER_MONGOLIA_SPLIT_2025[zone]
            ws.cell(row=row, column=2, value=year)
            ws.cell(row=row, column=3, value=solar_ele_gwh)
            ws.cell(row=row, column=4, value=float(m["value_factor_numerator"]))
            ws.cell(row=row, column=5, value=float(m["value_factor_denominator"]))
            ws.cell(row=row, column=6, value=float(m["value_factor"]))
            ws.cell(row=row, column=7, value=float(m["solar_penetration"]))
            ws.cell(row=row, column=8, value=float(m["solar_curtailment_rate"]))
            ws.cell(row=row, column=9, value=float(m["solar_capacity_factor"]))

    wb.save(XLSX_PATH)


def main() -> None:
    ap = argparse.ArgumentParser(description="Fill solar_value_dataset.xlsx from dispatch + price series.")
    ap.add_argument(
        "--mapped-csv",
        action="store_true",
        help="Use dispatch_segmented *_mapped.csv prices instead of planning postnetwork LMPs (legacy).",
    )
    args = ap.parse_args()
    price_source = "mapped_csv" if args.mapped_csv else "planning_marginal"

    metrics_by_year: dict[int, pd.DataFrame] = {}
    cap_compare_all: list[pd.DataFrame] = []
    for year in TARGET_YEARS:
        try:
            metrics, cap_compare = _compute_metrics_for_year(
                year,
                adjust_capacity=(year == 2025),
                price_source=price_source,
            )
        except FileNotFoundError as e:
            print(f"Skip {year}: {e}")
            continue
        metrics_by_year[year] = metrics
        cap_compare_all.append(cap_compare)

    if not metrics_by_year:
        raise RuntimeError(
            "No years were processed. Check dispatch_segmented .nc files and, when using default pricing, "
            "matching planning postnetworks under postnetworks/ (or pass --mapped-csv with CSV exports)."
        )

    if cap_compare_all:
        pd.concat(cap_compare_all, ignore_index=True).to_csv(CAP_COMPARE_PATH, index=False)

    _write_workbook(metrics_by_year)
    print(f"Price source: {price_source}")
    print(f"Filled years: {sorted(metrics_by_year.keys())}")
    print(f"Target years: {TARGET_YEARS}")
    print(f"Backup created: {BACKUP_PATH}")


if __name__ == "__main__":
    main()
